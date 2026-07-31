"""
Aplikasi Penarikan Kesimpulan Pemeriksaan Kepatuhan — Streamlit (single-file)
==============================================================================
Menjalankan:
    pip install streamlit pandas openpyxl
    streamlit run app.py

Struktur file ini (digabung dari beberapa modul supaya cukup 1 file):
  1) ENGINE   - perhitungan murni (bobot -> skor -> kesimpulan)
  2) EXPORTER - generator Lampiran 6.1-6.4 (satu file .xlsx, formula hidup)
  3) SEED     - data contoh dari workbook sumber (total = 1,66)
  4) APP (UI) - antarmuka Streamlit
==============================================================================
"""

from __future__ import annotations
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional
import io

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# 1) ENGINE
# ==============================================================================
def xround(value: float, digits: int = 2) -> float:
    """ROUND ala Excel (half away from zero), bukan banker's rounding Python.
    Contoh: xround(0.125, 2) -> 0.13 (Python round -> 0.12)."""
    if value is None:
        return None
    q = Decimal(10) ** -digits
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP))

# Nama parameter pembobotan temuan (Lamp 6.3). Bisa <4; rata-rata dibagi
# JUMLAH PARAMETER YANG AKTIF, bukan selalu 4 (koreksi atas hardcode /4 di Excel).
PARAMETER_TEMUAN = ["Nilai", "Dampak", "Sensitivitas", "Fraud"]


@dataclass
class Config:
    batas_penyimpangan: float = 0.3      # skala penyimpangan yang ditoleransi (default 0,3 = 10%)
    ambang_sesuai: float = 1.3           # < ini -> SESUAI
    ambang_tidak_sesuai: float = 1.8     # > ini -> TIDAK SESUAI


# ----------------------------------------------------------------------
# Fungsi-fungsi murni per baris
# ----------------------------------------------------------------------

def hitung_skala(mode: str, skala_langsung, params: dict) -> Optional[float]:
    """Kembalikan skala temuan 1..4, atau None kalau input belum lengkap."""
    if str(mode).lower().startswith("langsung"):
        try:
            v = float(skala_langsung)
        except (TypeError, ValueError):
            return None
        return v
    # mode parameter -> rata-rata parameter yang terisi (>0)
    nilai = [float(v) for v in params.values() if v is not None and str(v) != "" and float(v) > 0]
    if not nilai:
        return None
    return xround(sum(nilai) / len(nilai), 3)


def hitung_baris(bobot_aspek: float, bobot_kriteria: float, skala: Optional[float],
                 cfg: Config) -> dict:
    nilai_tertimbang = xround(bobot_aspek * bobot_kriteria, 3)
    if skala is None:
        return {
            "nilai_tertimbang": nilai_tertimbang,
            "skala": None, "skor_temuan": None, "skor_penyimpangan": None,
            "batas": None, "selisih": None, "simpulan_sub": "-",
        }
    skor_temuan = xround(nilai_tertimbang * skala, 2)
    skor_penyimpangan = xround(skor_temuan - nilai_tertimbang, 3)
    batas = xround(cfg.batas_penyimpangan * nilai_tertimbang, 3)
    selisih = xround(skor_penyimpangan - batas, 3)
    simpulan = ("MENYIMPANG SECARA MATERIAL" if selisih >= 0
                else "TIDAK MENYIMPANG SECARA MATERIAL")
    return {
        "nilai_tertimbang": nilai_tertimbang,
        "skala": skala,
        "skor_temuan": skor_temuan,
        "skor_penyimpangan": skor_penyimpangan,
        "batas": batas,
        "selisih": selisih,
        "simpulan_sub": simpulan,
    }


def kategori_kesimpulan(total_skor: float, cfg: Config) -> str:
    if total_skor < cfg.ambang_sesuai:
        return "SESUAI"
    if total_skor > cfg.ambang_tidak_sesuai:
        return "TIDAK SESUAI"
    return "SESUAI DENGAN PENGECUALIAN"


# ----------------------------------------------------------------------
# Hitung seluruh matriks dari dua DataFrame (aspek & sub-aspek)
# ----------------------------------------------------------------------

def hitung_matriks(df_aspek: pd.DataFrame, df_sub: pd.DataFrame, cfg: Config):
    """
    df_aspek kolom : kode, nama, bobot_aspek
    df_sub   kolom : kode_aspek, no, kriteria, bobot_kriteria, temuan,
                     mode, skala_langsung, Nilai, Dampak, Sensitivitas, Fraud
    Return: (df_hasil, ringkasan_dict)
    """
    bobot_map = {str(r["kode"]).strip(): float(r["bobot_aspek"] or 0)
                 for _, r in df_aspek.iterrows() if str(r.get("kode", "")).strip()}
    nama_map = {str(r["kode"]).strip(): str(r.get("nama", "") or "")
                for _, r in df_aspek.iterrows() if str(r.get("kode", "")).strip()}

    rows = []
    for _, r in df_sub.iterrows():
        kode = str(r.get("kode_aspek", "") or "").strip()
        if not kode:
            continue
        bobot_aspek = bobot_map.get(kode, 0.0)
        bobot_kriteria = float(r.get("bobot_kriteria") or 0)
        params = {p: r.get(p) for p in PARAMETER_TEMUAN}
        skala = hitung_skala(r.get("mode", "Parameter"), r.get("skala_langsung"), params)
        calc = hitung_baris(bobot_aspek, bobot_kriteria, skala, cfg)
        rows.append({
            "Kode Aspek": kode,
            "Aspek": nama_map.get(kode, ""),
            "Bobot Aspek": bobot_aspek,
            "No": r.get("no", ""),
            "Sub Aspek / Kriteria": r.get("kriteria", ""),
            "Bobot Kriteria": bobot_kriteria,
            "Nilai Tertimbang": calc["nilai_tertimbang"],
            "Temuan": r.get("temuan", ""),
            "Skala Temuan": calc["skala"],
            "Skor Temuan": calc["skor_temuan"],
            "Skor Penyimpangan": calc["skor_penyimpangan"],
            "Batas Penyimpangan": calc["batas"],
            "Selisih Penyimpangan": calc["selisih"],
            "Kesimpulan Sub-Aspek": calc["simpulan_sub"],
        })

    df_hasil = pd.DataFrame(rows)
    total_skor = xround(df_hasil["Skor Temuan"].fillna(0).sum(), 2) if not df_hasil.empty else 0.0
    kategori = kategori_kesimpulan(total_skor, cfg)

    ringkasan = {
        "total_skor": total_skor,
        "skor_penyimpangan": round(total_skor - 1, 3),
        "total_selisih": round(total_skor - 1 - cfg.batas_penyimpangan, 3),
        "kategori": kategori,
        "n_sub": len(df_hasil),
        "n_menyimpang": int((df_hasil["Kesimpulan Sub-Aspek"]
                             == "MENYIMPANG SECARA MATERIAL").sum()) if not df_hasil.empty else 0,
    }
    return df_hasil, ringkasan


# ----------------------------------------------------------------------
# Validasi bobot
# ----------------------------------------------------------------------

def validasi(df_aspek: pd.DataFrame, df_sub: pd.DataFrame, tol: float = 0.001):
    """Return list of (level, pesan). level: 'ok' | 'warn' | 'error'."""
    pesan = []
    total_aspek = df_aspek["bobot_aspek"].fillna(0).astype(float).sum()
    if abs(total_aspek - 1) <= tol:
        pesan.append(("ok", f"Sigma bobot aspek = {total_aspek:.3f} (valid)"))
    else:
        pesan.append(("error", f"Sigma bobot aspek = {total_aspek:.3f}, seharusnya 1,000"))

    for kode in df_aspek["kode"].dropna().astype(str).str.strip():
        if not kode:
            continue
        sub = df_sub[df_sub["kode_aspek"].astype(str).str.strip() == kode]
        if sub.empty:
            continue
        s = sub["bobot_kriteria"].fillna(0).astype(float).sum()
        if abs(s - 1) <= tol:
            pesan.append(("ok", f"Aspek {kode}: Sigma bobot kriteria = {s:.3f} (valid)"))
        else:
            pesan.append(("error", f"Aspek {kode}: Sigma bobot kriteria = {s:.3f}, seharusnya 1,000"))
    return pesan


# ----------------------------------------------------------------------
# Ekspor ke Excel (meniru layout Lampiran 6.4)
# ----------------------------------------------------------------------

def ekspor_excel(metadata: dict, df_hasil: pd.DataFrame, ringkasan: dict,
                 cfg: Config, judgment: str = "") -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriks Kesimpulan"

    bold = Font(name="Arial", bold=True, size=10)
    reg = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_menyimpang = PatternFill("solid", fgColor="FCE4E4")

    r = 1
    ws.cell(r, 1, "MATRIKS PENYIMPULAN PEMERIKSAAN KEPATUHAN").font = Font(
        name="Arial", bold=True, size=13)
    r += 2
    for label, key in [("Tujuan Pemeriksaan", "tujuan"), ("Subject Matter", "subject_matter"),
                       ("Lingkup", "lingkup"), ("Total Kontrak", "total_kontrak"),
                       ("Sampling", "sampling")]:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 2, ": " + str(metadata.get(key, ""))).font = reg
        r += 1
    r += 1

    headers = ["Kode Aspek", "Aspek", "Bobot Aspek", "No", "Sub Aspek / Kriteria",
               "Bobot Kriteria", "Nilai Tertimbang", "Temuan", "Skala Temuan",
               "Skor Temuan", "Skor Penyimpangan", "Batas Penyimpangan",
               "Selisih Penyimpangan", "Kesimpulan Sub-Aspek"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    header_row = r
    r += 1

    for _, row in df_hasil.iterrows():
        menyimpang = row["Kesimpulan Sub-Aspek"] == "MENYIMPANG SECARA MATERIAL"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, row[h])
            cell.font = reg
            cell.alignment = left if h in ("Aspek", "Sub Aspek / Kriteria", "Temuan",
                                           "Kesimpulan Sub-Aspek") else center
            cell.border = border
            if menyimpang:
                cell.fill = fill_menyimpang
        r += 1

    # baris total
    ws.cell(r, 5, "TOTAL SKOR TEMUAN").font = bold
    ws.cell(r, 10, ringkasan["total_skor"]).font = bold
    ws.cell(r, 10).alignment = center
    r += 2

    ws.cell(r, 1, "KESIMPULAN").font = Font(name="Arial", bold=True, size=12)
    r += 1
    kesimpulan_txt = (f'Total Skor Temuan {ringkasan["total_skor"]} '
                      f'-> kategori: {ringkasan["kategori"]}')
    ws.cell(r, 1, kesimpulan_txt).font = bold
    r += 1
    ws.cell(r, 1, f'Jumlah sub-aspek menyimpang material: {ringkasan["n_menyimpang"]} '
                  f'dari {ringkasan["n_sub"]}').font = reg
    r += 2
    if judgment.strip():
        ws.cell(r, 1, "Judgment Pervasiveness / Catatan Pemeriksa:").font = bold
        r += 1
        ws.cell(r, 1, judgment).font = reg
        ws.cell(r, 1).alignment = left
        r += 2
    ws.cell(r, 1, f"Parameter: batas penyimpangan={cfg.batas_penyimpangan}; "
                  f"ambang SESUAI<{cfg.ambang_sesuai}; "
                  f"TIDAK SESUAI>{cfg.ambang_tidak_sesuai}").font = Font(
        name="Arial", size=8, italic=True)

    widths = [11, 26, 9, 5, 34, 9, 11, 30, 9, 9, 12, 11, 12, 26]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================================
# 2) EXPORTER — Lampiran 6.1 s.d 6.4 (formula hidup)
# ==============================================================================
# ---- gaya ----
FONT = "Arial"
_reg = Font(name=FONT, size=10)
_bold = Font(name=FONT, bold=True, size=10)
_title = Font(name=FONT, bold=True, size=13)
_subhdr = Font(name=FONT, bold=True, size=11)
_small = Font(name=FONT, size=8, italic=True)
_hdrfont = Font(name=FONT, bold=True, size=10, color="FFFFFF")
_lamp = Font(name=FONT, bold=True, size=10, color="1F4E79")
_hdrfill = PatternFill("solid", fgColor="1F4E79")
_subfill = PatternFill("solid", fgColor="DDEBF7")
_menyimpang = PatternFill("solid", fgColor="FCE4E4")
_wrapL = Alignment(horizontal="left", vertical="center", wrap_text=True)
_wrapC = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="BFBFBF")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _put(ws, r, c, val, font=_reg, align=_wrapL, border=False, fill=None, numfmt=None):
    cell = ws.cell(r, c, val)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = _border
    if fill:
        cell.fill = fill
    if numfmt:
        cell.number_format = numfmt
    return cell


def _quote(sheet):
    return f"'{sheet}'"


def build_lampiran_workbook(metadata: dict, df_aspek, df_sub, cfg: Config,
                            judgment: str = "") -> bytes:
    wb = openpyxl.Workbook()
    S61 = "Lamp 6.1 Pembobotan Aspek"
    S62 = "Lamp 6.2 Pembobotan Sub aspek"
    S63 = "Lamp 6.3 Pembobotan Temuan"
    S64 = "Lamp 6.4 Kesimpulan"

    ws61 = wb.active; ws61.title = S61
    ws62 = wb.create_sheet(S62)
    ws63 = wb.create_sheet(S63)
    ws64 = wb.create_sheet(S64)

    # urutan aspek & sub
    aspek_list = [dict(kode=str(r["kode"]).strip(), nama=str(r.get("nama", "") or ""),
                       bobot=float(r.get("bobot_aspek") or 0),
                       pert=str(r.get("pertimbangan", "") or ""))
                  for _, r in df_aspek.iterrows() if str(r.get("kode", "")).strip()]
    subs_by_aspek = {}
    for a in aspek_list:
        rows = df_sub[df_sub["kode_aspek"].astype(str).str.strip() == a["kode"]]
        subs_by_aspek[a["kode"]] = list(rows.to_dict("records"))

    # peta koordinat utk cross-ref
    aspek_name_cell = {}   # kode -> "C6"
    aspek_bobot_cell = {}  # kode -> "F6"
    sub_bobot_cell = {}    # (kode, idx) -> "D15"  (bobot kriteria di 6.2)
    sub_rata_cell = {}     # (kode, idx) -> "K9"   (rata-rata skala di 6.3)

    # ================= LAMP 6.1 =================
    _put(ws61, 1, 7, "Lampiran 6.1", _lamp, _wrapC)
    ws61.merge_cells("B3:G3"); _put(ws61, 3, 2, "Pembobotan Aspek", _subhdr, _wrapC)
    _put(ws61, 5, 3, "Aspek", _hdrfont, _wrapC, True, _hdrfill)
    ws61.merge_cells("C5:E5")
    for cc in range(4, 6):
        ws61.cell(5, cc).fill = _hdrfill; ws61.cell(5, cc).border = _border
    _put(ws61, 5, 2, "No", _hdrfont, _wrapC, True, _hdrfill)
    _put(ws61, 5, 6, "Bobot", _hdrfont, _wrapC, True, _hdrfill)
    _put(ws61, 5, 7, "Pertimbangan", _hdrfont, _wrapC, True, _hdrfill)
    r = 6
    letters = "abcdefghijklmnop"
    for i, a in enumerate(aspek_list):
        _put(ws61, r, 2, f"{letters[i]}.", _reg, _wrapC, True)
        ws61.merge_cells(f"C{r}:E{r}")
        _put(ws61, r, 3, a["nama"], _reg, _wrapL, True)
        for cc in range(4, 6):
            ws61.cell(r, cc).border = _border
        _put(ws61, r, 6, a["bobot"], _reg, _wrapC, True, numfmt="0.000")
        _put(ws61, r, 7, a["pert"], _reg, _wrapL, True)
        aspek_name_cell[a["kode"]] = f"C{r}"
        aspek_bobot_cell[a["kode"]] = f"F{r}"
        r += 1
    _put(ws61, r, 3, "Jumlah", _bold, _wrapC, True)
    first_b, last_b = 6, r - 1
    _put(ws61, r, 6, f"=SUM(F{first_b}:F{last_b})", _bold, _wrapC, True, numfmt="0.000")
    for w, col in zip([5, 4, 20, 6, 6, 10, 55], "ABCDEFG"):
        ws61.column_dimensions[col].width = w

    # ================= LAMP 6.2 =================
    _put(ws62, 1, 5, "Lampiran 6.2", _lamp, _wrapC)
    ws62.merge_cells("A2:E2"); _put(ws62, 2, 1, "PEMBOBOTAN SUB ASPEK / KRITERIA", _subhdr, _wrapC)
    _put(ws62, 4, 1, "Masing-masing aspek di-breakdown menjadi sub aspek yang didasarkan "
                     "pada kriteria utama yang relevan dengan tujuan pemeriksaan.", _reg, _wrapL)
    ws62.merge_cells("A4:E4")
    hdr = 6
    for c, h in zip(range(1, 6), ["NO", "ASPEK", "SUB ASPEK/KRITERIA", "BOBOT KRITERIA", "PERTIMBANGAN"]):
        _put(ws62, hdr, c, h, _hdrfont, _wrapC, True, _hdrfill)
    r = hdr + 1
    for a in aspek_list:
        subs = subs_by_aspek[a["kode"]]
        start = r
        for j, s in enumerate(subs):
            if j == 0:
                _put(ws62, r, 1, a["kode"], _reg, _wrapC, True)
                _put(ws62, r, 2, f"={_quote(S61)}!{aspek_name_cell[a['kode']]}", _reg, _wrapL, True)
            else:
                _put(ws62, r, 1, "", _reg, _wrapC, True)
                _put(ws62, r, 2, "", _reg, _wrapL, True)
            _put(ws62, r, 3, str(s.get("kriteria", "") or ""), _reg, _wrapL, True)
            _put(ws62, r, 4, float(s.get("bobot_kriteria") or 0), _reg, _wrapC, True, numfmt="0.000")
            _put(ws62, r, 5, str(s.get("pertimbangan", "") or ""), _reg, _wrapL, True)
            sub_bobot_cell[(a["kode"], j)] = f"D{r}"
            r += 1
        _put(ws62, r, 2, "Subjumlah", _bold, _wrapC, True, _subfill)
        _put(ws62, r, 1, "", _reg, _wrapC, True, _subfill)
        _put(ws62, r, 3, "", _reg, _wrapL, True, _subfill)
        _put(ws62, r, 4, f"=SUM(D{start}:D{r-1})", _bold, _wrapC, True, _subfill, "0.000")
        _put(ws62, r, 5, "", _reg, _wrapL, True, _subfill)
        r += 1
    for w, col in zip([5, 20, 40, 12, 45], "ABCDE"):
        ws62.column_dimensions[col].width = w

    # ================= LAMP 6.3 =================
    _put(ws63, 1, 11, "Lampiran 6.3", _lamp, _wrapC)
    _put(ws63, 2, 2, "Pembobotan Temuan", _subhdr, _wrapL)
    _put(ws63, 4, 2, "Parameter pertimbangan: Nilai, Dampak, Sensitivitas/harapan publik, "
                     "Indikasi Fraud. Skala 1 (rendah) s.d 4 (tinggi).", _reg, _wrapL)
    _put(ws63, 5, 2, "Rata-rata = jumlah parameter dibagi banyaknya parameter yang dipakai.", _small, _wrapL)
    hdr = 7
    heads = ["Sub aspek/kriteria", "Temuan", "Nilai", "Dampak",
             "Sensitivitas/ harapan publik", "Indikasi Fraud", "Jumlah", "Rata Rata*"]
    for c, h in zip(range(4, 12), heads):   # D..K
        _put(ws63, hdr, c, h, _hdrfont, _wrapC, True, _hdrfill)
    r = hdr + 1
    param_cols = {"Nilai": 6, "Dampak": 7, "Sensitivitas": 8, "Fraud": 9}  # F,G,H,I
    for a in aspek_list:
        for j, s in enumerate(subs_by_aspek[a["kode"]]):
            _put(ws63, r, 4, str(s.get("kriteria", "") or ""), _reg, _wrapL, True)
            _put(ws63, r, 5, str(s.get("temuan", "") or ""), _reg, _wrapL, True)
            mode = str(s.get("mode", "Parameter"))
            present = []
            if mode.lower().startswith("langsung"):
                for p, cc in param_cols.items():
                    _put(ws63, r, cc, "", _reg, _wrapC, True)
                _put(ws63, r, 10, "", _reg, _wrapC, True)  # J
                skl = s.get("skala_langsung")
                skl = float(skl) if skl not in (None, "") else 0
                _put(ws63, r, 11, skl, _reg, _wrapC, True, numfmt="0.000")  # K langsung
            else:
                for p, cc in param_cols.items():
                    v = s.get(p)
                    if v not in (None, "") and float(v) > 0:
                        _put(ws63, r, cc, float(v), _reg, _wrapC, True)
                        present.append(get_column_letter(cc) + str(r))
                    else:
                        _put(ws63, r, cc, "", _reg, _wrapC, True)
                if present:
                    _put(ws63, r, 10, "=" + "+".join(present), _reg, _wrapC, True)  # Jumlah
                    _put(ws63, r, 11, f"=ROUND(J{r}/{len(present)},3)", _reg, _wrapC, True, numfmt="0.000")
                else:
                    _put(ws63, r, 10, "", _reg, _wrapC, True)
                    _put(ws63, r, 11, 0, _reg, _wrapC, True, numfmt="0.000")
            sub_rata_cell[(a["kode"], j)] = f"K{r}"
            r += 1
    _put(ws63, r + 1, 4, "*Pembagi = banyaknya parameter yang digunakan (bisa < 4).", _small, _wrapL)
    for w, col in zip([3, 3, 3, 32, 26, 7, 7, 12, 7, 8, 10], "ABCDEFGHIJK"):
        ws63.column_dimensions[col].width = w

    # ================= LAMP 6.4 =================
    _put(ws64, 1, 16, "Lampiran 6.4", _lamp, _wrapC)
    _put(ws64, 1, 1, "MATRIKS PENYIMPULAN PEMERIKSAAN KEPATUHAN", _title, _wrapL)
    r = 3
    for label, key in [("TUJUAN PEMERIKSAAN", "tujuan"), ("SUBJECT MATTER", "subject_matter"),
                       ("LINGKUP", "lingkup"), ("TOTAL KONTRAK", "total_kontrak"),
                       ("SAMPLING", "sampling")]:
        _put(ws64, r, 1, label, _bold, _wrapL)
        _put(ws64, r, 3, ": " + str(metadata.get(key, "")), _reg, _wrapL)
        r += 1
    r += 1

    heads = ["NO", "ASPEK", "PEMBOBOTAN", "SUB ASPEK/KRITERIA", "BOBOT KRITERIA",
             "NILAI TERTIMBANG", "TEMUAN", "SKALA TEMUAN", "SKOR TEMUAN",
             "SKOR PENYIMPANGAN TERTIMBANG", "BATAS PENYIMPANGAN (tertimbang)",
             "SELISIH PENYIMPANGAN", "KESIMPULAN PER SUBASPEK"]
    # kolom: A=NO(kode), B=Aspek, D=Pembobotan(bobot aspek), E=No sub, F=Kriteria,
    #        G=Bobot kriteria, H=Nilai tertimbang, I=Temuan, K=Skala, L=Skor,
    #        M=Skor peny, N=Batas, O=Selisih, P=Kesimpulan
    colmap = {"NO": 1, "ASPEK": 2, "PEMBOBOTAN": 4, "SUB ASPEK/KRITERIA": 6,
              "BOBOT KRITERIA": 7, "NILAI TERTIMBANG": 8, "TEMUAN": 9, "SKALA TEMUAN": 11,
              "SKOR TEMUAN": 12, "SKOR PENYIMPANGAN TERTIMBANG": 13,
              "BATAS PENYIMPANGAN (tertimbang)": 14, "SELISIH PENYIMPANGAN": 15,
              "KESIMPULAN PER SUBASPEK": 16}
    hdr = r
    for h in heads:
        _put(ws64, hdr, colmap[h], h, _hdrfont, _wrapC, True, _hdrfill)
    # kolom kosong tetap diberi border header
    for c in range(1, 17):
        ws64.cell(hdr, c).fill = _hdrfill; ws64.cell(hdr, c).border = _border
    numrow = hdr + 1
    for c, t in [(1, "(1)"), (4, "(2)"), (6, "(3)"), (7, "(4)"), (8, "(5)=(2)*(4)"),
                 (9, "(6)"), (11, "(7)"), (12, "(8)=(5)*(7)"), (13, "(9)=(8)-(5)"),
                 (14, "(10)=bts*(5)"), (15, "(11)=(9)-(10)"), (16, "(12)")]:
        _put(ws64, numrow, c, t, _small, _wrapC, True)
    for c in range(1, 17):
        ws64.cell(numrow, c).border = _border

    r = numrow + 1
    data_start = r
    batas = cfg.batas_penyimpangan
    for a in aspek_list:
        subs = subs_by_aspek[a["kode"]]
        group_first = r
        for j, s in enumerate(subs):
            if j == 0:
                _put(ws64, r, 1, a["kode"], _reg, _wrapC, True)
                _put(ws64, r, 2, f"={_quote(S61)}!{aspek_name_cell[a['kode']]}", _reg, _wrapL, True)
                _put(ws64, r, 4, f"={_quote(S61)}!{aspek_bobot_cell[a['kode']]}", _reg, _wrapC, True, numfmt="0.000")
            else:
                _put(ws64, r, 1, "", _reg, _wrapC, True)
                _put(ws64, r, 2, "", _reg, _wrapL, True)
                _put(ws64, r, 4, "", _reg, _wrapC, True)
            _put(ws64, r, 5, s.get("no", j + 1), _reg, _wrapC, True)              # No sub-aspek
            _put(ws64, r, 6, str(s.get("kriteria", "") or ""), _reg, _wrapL, True)  # Kriteria
            _put(ws64, r, 7, f"={_quote(S62)}!{sub_bobot_cell[(a['kode'], j)]}", _reg, _wrapC, True, numfmt="0.000")
            bobot_aspek_ref = f"$D${group_first}"
            _put(ws64, r, 8, f"=ROUND({bobot_aspek_ref}*G{r},3)", _reg, _wrapC, True, numfmt="0.000")
            _put(ws64, r, 9, str(s.get("temuan", "") or ""), _reg, _wrapL, True)
            _put(ws64, r, 11, f"={_quote(S63)}!{sub_rata_cell[(a['kode'], j)]}", _reg, _wrapC, True, numfmt="0.000")
            _put(ws64, r, 12, f"=ROUND(H{r}*K{r},2)", _reg, _wrapC, True, numfmt="0.00")
            _put(ws64, r, 13, f"=L{r}-H{r}", _reg, _wrapC, True, numfmt="0.000")
            _put(ws64, r, 14, f"=ROUND({batas}*H{r},3)", _reg, _wrapC, True, numfmt="0.000")
            _put(ws64, r, 15, f"=M{r}-N{r}", _reg, _wrapC, True, numfmt="0.000")
            _put(ws64, r, 16,
                 f'=IF(O{r}<0,"TIDAK MENYIMPANG SECARA MATERIAL","MENYIMPANG SECARA MATERIAL")',
                 _reg, _wrapL, True)
            r += 1
    data_end = r - 1

    # baris jumlah
    _put(ws64, r, 4, f"=SUM(D{data_start}:D{data_end})", _bold, _wrapC, True, numfmt="0.000")
    _put(ws64, r, 6, "Jumlah", _bold, _wrapL, True)
    _put(ws64, r, 8, f"=SUM(H{data_start}:H{data_end})", _bold, _wrapC, True, numfmt="0.000")
    _put(ws64, r, 12, f"=ROUND(SUM(L{data_start}:L{data_end}),2)", _bold, _wrapC, True, numfmt="0.00")
    total_cell = f"L{r}"
    r += 3

    _put(ws64, r, 2, "HAL-HAL YANG DISEPAKATI", _bold, _wrapL); r += 1
    _put(ws64, r, 2, "Skala/Skor Temuan", _reg, _wrapL); _put(ws64, r, 4, "1 s.d 4", _reg, _wrapC); r += 1
    _put(ws64, r, 2, "Batas Penyimpangan", _reg, _wrapL); _put(ws64, r, 4, batas, _reg, _wrapC, numfmt="0.000"); r += 1
    _put(ws64, r, 2, f"Ambang SESUAI < {cfg.ambang_sesuai}; TIDAK SESUAI > {cfg.ambang_tidak_sesuai}", _reg, _wrapL); r += 2

    _put(ws64, r, 2, "KESIMPULAN", Font(name=FONT, bold=True, size=12), _wrapL); r += 1
    kes = (f'=CONCATENATE("Total Skor Temuan ",{total_cell}," berarti masuk kategori: ",'
           f'IF({total_cell}<{cfg.ambang_sesuai},"SESUAI",'
           f'IF({total_cell}>{cfg.ambang_tidak_sesuai},"TIDAK SESUAI","SESUAI DENGAN PENGECUALIAN")))')
    _put(ws64, r, 2, kes, _bold, _wrapL); r += 2
    if judgment.strip():
        _put(ws64, r, 2, "Judgment Pervasiveness / Catatan Pemeriksa:", _bold, _wrapL); r += 1
        _put(ws64, r, 2, judgment, _reg, _wrapL); r += 1

    widths = [5, 22, 4, 10, 5, 30, 9, 11, 26, 9, 11, 14, 13, 13, 13, 26]
    for i, w in enumerate(widths, 1):
        ws64.column_dimensions[get_column_letter(i)].width = w
    ws64.column_dimensions["C"].width = 6

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================================
# 3) SEED DATA — contoh dari workbook sumber
# ==============================================================================
METADATA_DEFAULT = {
    "tujuan": "Memberikan simpulan apakah kegiatan pelaksanaan belanja telah sesuai kriteria",
    "subject_matter": "Pelaksanaan belanja terkait ...",
    "lingkup": "Pelaksanaan belanja infrastruktur gedung, jalan, dan jembatan",
    "total_kontrak": "Rp100.000.000.000",
    "sampling": "Rp50.000.000.000 (50% populasi)",
}

ASPEK_SEED = pd.DataFrame([
    {"kode": "A", "nama": "Persiapan Pengadaan", "bobot_aspek": 0.2,
     "pertimbangan": "- Ketentuan terkait persiapan pengadaan telah dipahami oleh PPK.\n"
                     "- Persiapan pengadaan dilakukan sebelum penetapan pagu anggaran.\n"
                     "- Diatur jelas dalam Permen PUPR 14/2020 dan Peraturan LKPP No 12/2021."},
    {"kode": "B", "nama": "Pemilihan Penyedia", "bobot_aspek": 0.2,
     "pertimbangan": "- Pemilihan penyedia menentukan keberhasilan pekerjaan.\n"
                     "- Adanya potensi persekongkolan."},
    {"kode": "C", "nama": "Pelaksanaan Kontrak dan Serah Terima", "bobot_aspek": 0.5,
     "pertimbangan": "Tahap pelaksanaan menentukan penyelesaian pekerjaan tepat waktu, "
                     "kuantitas, dan kualitas."},
    {"kode": "D", "nama": "Pembayaran Prestasi Pekerjaan", "bobot_aspek": 0.1,
     "pertimbangan": "Pembayaran bersifat seragam dan melalui beberapa verifikasi."},
])

# kode_aspek, no, kriteria, bobot_kriteria, temuan, mode, skala_langsung, Nilai, Dampak, Sensitivitas, Fraud
_sub = [
    ("A", 1, "PPK telah menyusun dan menetapkan HPS/Referensi Harga sesuai ketentuan", 0.3, "", 3, 4, 3, 3),
    ("A", 2, "PPK telah menetapkan rancangan kontrak yang sesuai dengan ketentuan", 0.2, "", 1, 1, 1, 1),
    ("A", 3, "PPK telah menetapkan spesifikasi teknis/KAK", 0.3, "", 3, 3, 3, 3),
    ("A", 4, "PPK telah menetapkan uang muka dan jaminan-jaminan sesuai ketentuan", 0.2, "", 1, 1, 1, 1),
    ("B", 1, "Metode pemilihan penyedia sudah tepat", 0.4, "", 1, 1, 1, 1),
    ("B", 2, "Pelaksanaan evaluasi atas dokumen penawaran peserta sesuai ketentuan", 0.6, "", 3, 4, 3, 3),
    ("C", 1, "Administrasi pelaksanaan paket pekerjaan sudah lengkap dan sesuai", 0.1, "", 1, 1, 1, 1),
    ("C", 2, "Kuantitas hasil pekerjaan sudah sesuai kontrak", 0.35, "", 1.5, 1.5, 1, 1),
    ("C", 3, "Kualitas hasil pekerjaan sudah sesuai kontrak", 0.35, "", 1, 1, 1, 1),
    ("C", 4, "Pelaksanaan pekerjaan selesai tepat waktu", 0.1, "", 1, 1, 1, 1),
    ("C", 5, "Pengenaan denda telah sesuai ketentuan", 0.1, "", 1, 1, 1, 1),
    ("D", 1, "Pembayaran dilengkapi dengan bukti yang lengkap dan sah", 0.5, "", 2, 3, 2, 3),
    ("D", 2, "Realisasi pembayaran termin telah sesuai dengan prestasi pekerjaan", 0.5, "", 1, 1, 1, 1),
]

_PERT_SUB = {"A|1": "- Penyusunan dan penetapan HPS merupakan hal yang penting.\\n- HPS dapat dijadikan modus mark up berdampak besar.", "A|3": "Terdapat risiko penetapan spesifikasi teknis yang menguntungkan pihak tertentu.", "B|2": "Evaluasi penawaran mempunyai efek sangat signifikan dalam pemilihan penyedia.", "C|2": "Ketidaktepatan kuantitas/kualitas berdampak buruk: masa manfaat berkurang, keamanan konstruksi, kelebihan pembayaran."}

SUB_SEED = pd.DataFrame(
    [{"kode_aspek": k, "no": n, "kriteria": kr, "bobot_kriteria": b, "temuan": t,
      "pertimbangan": _PERT_SUB.get(f"{k}|{n}", ""),
      "mode": "Parameter", "skala_langsung": None,
      "Nilai": ni, "Dampak": da, "Sensitivitas": se, "Fraud": fr}
     for (k, n, kr, b, t, ni, da, se, fr) in _sub]
)


# ==============================================================================
# 4) APLIKASI STREAMLIT (UI)
# ==============================================================================
st.set_page_config(page_title="Matriks Penyimpulan Pemeriksaan Kepatuhan",
                   page_icon="📋", layout="wide")

# ---------------------------------------------------------------- state init
if "aspek" not in st.session_state:
    st.session_state.aspek = ASPEK_SEED.copy()
    st.session_state.sub = SUB_SEED.copy()
    st.session_state.meta = dict(METADATA_DEFAULT)
    st.session_state.judgment = ""

# ---------------------------------------------------------------- sidebar
with st.sidebar:
    st.header("⚙️ Pengaturan")
    batas = st.number_input("Batas penyimpangan (skala)", 0.0, 3.0,
                            value=0.3, step=0.05,
                            help="Default 0,3 (≈10%). Diproporsionalkan per sub-aspek: batas × nilai tertimbang.")
    amb_sesuai = st.number_input("Ambang SESUAI (< nilai ini)", 0.0, 4.0, value=1.3, step=0.1)
    amb_tidak = st.number_input("Ambang TIDAK SESUAI (> nilai ini)", 0.0, 4.0, value=1.8, step=0.1)
    cfg = Config(batas_penyimpangan=batas, ambang_sesuai=amb_sesuai,
                 ambang_tidak_sesuai=amb_tidak)

    st.divider()
    st.caption("Reset ke data contoh workbook:")
    if st.button("↺ Muat ulang data contoh", use_container_width=True):
        st.session_state.aspek = ASPEK_SEED.copy()
        st.session_state.sub = SUB_SEED.copy()
        st.session_state.meta = dict(METADATA_DEFAULT)
        st.session_state.judgment = ""
        st.rerun()
    if st.button("🗑️ Kosongkan semua", use_container_width=True):
        st.session_state.aspek = ASPEK_SEED.iloc[0:0].copy()
        st.session_state.sub = SUB_SEED.iloc[0:0].copy()
        st.rerun()

st.title("📋 Matriks Penyimpulan Pemeriksaan Kepatuhan")
st.caption("Replikasi Lampiran 6.1–6.4 · pembobotan Aspek → Sub-aspek → Temuan · rekalkulasi otomatis")

# ---------------------------------------------------------------- metadata
with st.expander("🧾 Identitas Pemeriksaan", expanded=False):
    m = st.session_state.meta
    c1, c2 = st.columns(2)
    m["tujuan"] = c1.text_area("Tujuan Pemeriksaan", m.get("tujuan", ""), height=70)
    m["subject_matter"] = c2.text_area("Subject Matter", m.get("subject_matter", ""), height=70)
    m["lingkup"] = c1.text_input("Lingkup", m.get("lingkup", ""))
    m["total_kontrak"] = c2.text_input("Total Kontrak", m.get("total_kontrak", ""))
    m["sampling"] = c1.text_input("Sampling", m.get("sampling", ""))

# ---------------------------------------------------------------- input aspek
st.subheader("1️⃣ Aspek & Bobot")
st.caption("Σ bobot aspek harus = 1,000. Tambah/hapus baris lewat tabel.")
aspek_ed = st.data_editor(
    st.session_state.aspek, num_rows="dynamic", use_container_width=True,
    key="ed_aspek",
    column_config={
        "kode": st.column_config.TextColumn("Kode", width="small", required=True),
        "nama": st.column_config.TextColumn("Nama Aspek", width="medium"),
        "bobot_aspek": st.column_config.NumberColumn("Bobot", min_value=0.0,
                                                     max_value=1.0, step=0.05, format="%.3f"),
        "pertimbangan": st.column_config.TextColumn("Pertimbangan (Lamp 6.1)", width="large"),
    })
st.session_state.aspek = aspek_ed
tot_aspek = pd.to_numeric(aspek_ed["bobot_aspek"], errors="coerce").fillna(0).sum()
(st.success if abs(tot_aspek - 1) <= 0.001 else st.error)(
    f"Σ bobot aspek = {tot_aspek:.3f}" + ("" if abs(tot_aspek - 1) <= 0.001 else "  (harus 1,000)"))

# ---------------------------------------------------------------- input sub-aspek
st.subheader("2️⃣ Sub-Aspek / Kriteria & Skala Temuan")
st.caption("Mode **Parameter**: isi Nilai/Dampak/Sensitivitas/Fraud (1–4), skala = rata-rata "
           "parameter terisi. Mode **Langsung**: pakai kolom *Skala Langsung* (1–4).")
kode_opsi = [str(k).strip() for k in aspek_ed["kode"].dropna().tolist() if str(k).strip()]
sub_ed = st.data_editor(
    st.session_state.sub, num_rows="dynamic", use_container_width=True,
    key="ed_sub",
    column_config={
        "kode_aspek": st.column_config.SelectboxColumn("Aspek", options=kode_opsi,
                                                       width="small", required=True),
        "no": st.column_config.NumberColumn("No", width="small", step=1),
        "kriteria": st.column_config.TextColumn("Sub-Aspek / Kriteria", width="large"),
        "bobot_kriteria": st.column_config.NumberColumn("Bobot", min_value=0.0,
                                                        max_value=1.0, step=0.05, format="%.3f"),
        "pertimbangan": st.column_config.TextColumn("Pertimbangan (Lamp 6.2)", width="medium"),
        "temuan": st.column_config.TextColumn("Temuan", width="medium"),
        "mode": st.column_config.SelectboxColumn("Mode", options=["Parameter", "Langsung"],
                                                 width="small"),
        "skala_langsung": st.column_config.NumberColumn("Skala Langsung", min_value=1.0,
                                                        max_value=4.0, step=0.25, format="%.2f"),
        "Nilai": st.column_config.NumberColumn("Nilai", min_value=1.0, max_value=4.0, step=0.5),
        "Dampak": st.column_config.NumberColumn("Dampak", min_value=1.0, max_value=4.0, step=0.5),
        "Sensitivitas": st.column_config.NumberColumn("Sensitiv.", min_value=1.0, max_value=4.0, step=0.5),
        "Fraud": st.column_config.NumberColumn("Fraud", min_value=1.0, max_value=4.0, step=0.5),
    })
st.session_state.sub = sub_ed

# validasi bobot kriteria per aspek
for lvl, msg in validasi(aspek_ed, sub_ed):
    if "kriteria" in msg and lvl == "error":
        st.warning(msg)

# ---------------------------------------------------------------- hitung
df_hasil, ring = hitung_matriks(aspek_ed, sub_ed, cfg)

st.subheader("3️⃣ Matriks Hasil (otomatis)")
if df_hasil.empty:
    st.info("Belum ada sub-aspek untuk dihitung.")
else:
    def _highlight(row):
        warna = "background-color:#fce4e4" if row["Kesimpulan Sub-Aspek"] == "MENYIMPANG SECARA MATERIAL" else ""
        return [warna] * len(row)
    st.dataframe(
        df_hasil.style.apply(_highlight, axis=1).format({
            "Bobot Aspek": "{:.3f}", "Bobot Kriteria": "{:.3f}",
            "Nilai Tertimbang": "{:.3f}", "Skala Temuan": "{:.3f}",
            "Skor Temuan": "{:.2f}", "Skor Penyimpangan": "{:.3f}",
            "Batas Penyimpangan": "{:.3f}", "Selisih Penyimpangan": "{:.3f}",
        }, na_rep="-"),
        use_container_width=True, height=460)

# ---------------------------------------------------------------- kesimpulan
st.subheader("4️⃣ Kesimpulan")
warna_kat = {"SESUAI": "green", "SESUAI DENGAN PENGECUALIAN": "orange",
             "TIDAK SESUAI": "red"}.get(ring["kategori"], "gray")
c1, c2, c3 = st.columns(3)
c1.metric("Total Skor Temuan", f'{ring["total_skor"]:.2f}')
c2.metric("Sub-aspek menyimpang", f'{ring["n_menyimpang"]} / {ring["n_sub"]}')
c3.metric("Total selisih penyimpangan", f'{ring["total_selisih"]:.3f}')
st.markdown(f"### Kategori: :{warna_kat}[**{ring['kategori']}**]")

if ring["kategori"] == "SESUAI DENGAN PENGECUALIAN":
    st.info("Kategori 'sesuai dengan pengecualian' memerlukan **judgment pervasiveness**: "
            "tentukan apakah temuan menyimpang terkonsentrasi pada aspek tertentu.")

st.session_state.judgment = st.text_area(
    "Judgment pervasiveness / catatan pemeriksa",
    st.session_state.judgment,
    placeholder="Mis. Penyimpangan hanya pada Aspek C dan tidak pervasive → "
                "'Sesuai dengan pengecualian pada Aspek C'.",
    height=90)

# ---------------------------------------------------------------- ekspor
st.subheader("5️⃣ Unduh Lampiran")
st.caption("Satu file Excel berisi 4 sheet — **Lampiran 6.1, 6.2, 6.3, 6.4** — dengan struktur "
           "dan formula hidup persis workbook asli. Nilai terhitung ulang otomatis di Excel.")
if df_hasil.empty:
    st.info("Isi data dulu untuk mengunduh lampiran.")
else:
    xlsx_full = build_lampiran_workbook(st.session_state.meta, aspek_ed, sub_ed, cfg,
                                        st.session_state.judgment)
    st.download_button("⬇️ Unduh Matriks_Kesimpulan.xlsx (Lampiran 6.1–6.4)",
                       data=xlsx_full, file_name="Matriks_Kesimpulan.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)
